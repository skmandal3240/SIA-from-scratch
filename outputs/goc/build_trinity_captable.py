#!/usr/bin/env python3
"""Project TRINITY — Cap Table x Valuation Scenario Workbook (xlsx).
Two sheets: (1) Cap table dilution path through Seed→Series C, (2) valuation
scenario matrix (base/bull/bear) with post-money, dilution %, and founder %.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

OUT = Path("/root/SIA-from-scratch/outputs/goc/TRINITY_CapTable_Valuation.xlsx")
NAVY = Font(color="1F1B5A", bold=True, size=12)

wb = Workbook()

# ---------------- Sheet 1: Cap Table -----------------
ws = wb.active
ws.title = "Cap Table"
ws.append(["Project TRINITY — Cap Table & Dilution Waterfall (illustrative)"])
ws["A1"].font = NAVY

rows = [
    ["Round", "Pre-money ₹Cr", "Invest ₹Cr", "Post-money ₹Cr", "New shares", "Founder%", "Investors%"],
    ["Founder", 0, 0, 0, 0, 1.00, 0.00],
    ["Seed (MVP)", 20, 5, 25, 2500, 0.80, 0.20],
    ["Series A", 114, 25, 139, 4397, 0.656, 0.344],
    ["Series B", 567, 100, 667, 5313, 0.558, 0.442],
    ["Series C", 1833, 500, 2333, 6837, 0.491, 0.509],
]
for r in rows:
    ws.append(r)
for cell in ws[1]:
    cell.font = Font(bold=True)

# ---------------- Sheet 2: Valuation Scenarios -----------------
ws3 = wb.create_sheet("Valuation Scenarios")
ws3.append(["Project TRINITY — Valuation Scenario Matrix (illustrative, ₹ Cr)"])
ws3["A1"].font = NAVY
ws3.append([""])

scenarios = [
    ["Scenario", "Trigger", "Y5 Revenue", "Exit multiple (comp)", "Exit value", "Investor ownership", "ROI"],
    ["Bear", "Pilots stall; no gov contract", "₹20", "4x", "₹80", "45%", "1.1x"],
    ["Base", "5 paid + 1 gov contract", "₹68", "6x", "₹408", "35%", "3.0x"],
    ["Bull", "National rollout; 2-3 gov contracts", "₹150", "8x", "₹1,200", "30%", "5.5x"],
    ["Teardown", "Direct strategic (sovereign asset)", "₹150", "12x", "₹1,800", "30%", "8.0x"],
]
for r in scenarios:
    ws3.append(r)
for cell in ws3[2]:
    cell.font = Font(bold=True)

# ---------------- Sheet 3: Round Structure -----------------
ws4 = wb.create_sheet("Round Structure")
ws4.append(["Project TRINITY — Funding Round Structure (illustrative)"])
ws4["A1"].font = NAVY
ws4.append(["Round", "Ask ₹Cr", "Dilution %", "ESOP carve %", "Post-money ₹Cr", "Founder post-round %"])
for r in [
    ["Seed (MVP)", 5, 15, 10, 25, 80],
    ["Series A", 25, 15, 12, 139, 65],
    ["Series B", 100, 12, 10, 667, 55],
    ["Series C", 500, 10, 8, 2333, 49],
]:
    ws4.append(r)
for cell in ws4[2]:
    cell.font = Font(bold=True)

wb.save(OUT)
print(f"SAVED {OUT}")