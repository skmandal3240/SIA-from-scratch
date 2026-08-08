#!/usr/bin/env python3
"""Project TRINITY — 3-statement financial model (xlsx). 10-year, illustrative."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

OUT = Path("/root/SIA-from-scratch/outputs/goc/TRINITY_Financial_Model.xlsx")
NAVY = Font(color="1F1B5A", bold=True, size=12)

years = list(range(2026, 2036))  # 10 years

revenue_lines = {
    "SIA Edge consumer": [0, 0.5, 2, 4, 8, 15, 25, 40, 50, 60],
    "SIA API (cloud tier)": [0, 0, 1, 3, 10, 20, 30, 45, 60, 75],
    "Enterprise subscriptions (SIA Pro)": [0, 0, 5, 12, 25, 45, 70, 150, 200, 260],
    "Government contracts": [0, 0, 3, 7, 15, 30, 50, 65, 80, 100],
    "SIA Cloud / managed services": [0, 0, 0, 2, 10, 25, 50, 80, 110, 140],
}
total_revenue = [round(sum(v[i] for v in revenue_lines.values()), 2) for i in range(10)]
COGS = [0.30, 0.35, 0.35, 0.30, 0.25, 0.25, 0.20, 0.20, 0.18, 0.18]
gross_profit = [round(total_revenue[i] * (1 - COGS[i]), 2) for i in range(10)]

opex = {
    "R&D / engineering": [0.8, 2.0, 4.0, 8.0, 15.0, 25.0, 35.0, 45.0, 60.0, 80.0],
    "Sales & marketing": [0.3, 1.0, 2.5, 6.0, 12.0, 20.0, 30.0, 40.0, 55.0, 70.0],
    "G&A": [0.3, 0.8, 1.5, 3.0, 5.0, 8.0, 12.0, 15.0, 20.0, 25.0],
    "GPU rented": [0.5, 2.0, 4.0, 8.0, 12.0, 15.0, 10.0, 8.0, 6.0, 5.0],
}
total_opex = [round(sum(v[i] for v in opex.values()), 2) for i in range(10)]
ebitda = [round(gross_profit[i] - total_opex[i], 2) for i in range(10)]
ebitda_margin = [round(ebitda[i] / total_revenue[i] * 100, 1) if total_revenue[i] else 0 for i in range(10)]

wb = Workbook()

# ---- Income Statement ----
ws = wb.active
ws.title = "Income Statement"
ws.append(["Project TRINITY — Income Statement (₹ crore, illustrative)"])
ws["A1"].font = NAVY
ws.append(["Line item"] + [f"{y}" for y in years])
def add(label, vals=None, fmt='₹#,##0.00'):
    if vals is None:
        ws.append([label])
        return
    ws.append([label] + [v for v in vals])
    for c in range(2, 12):
        ws.cell(row=ws.max_row, column=c).number_format = fmt
add("Revenue")
for name, vals in revenue_lines.items():
    add(f"  {name}", vals)
add("Total revenue", total_revenue)
add("COGS (subtotal)", [round(total_revenue[i] * COGS[i], 2) for i in range(10)])
add("Gross profit", gross_profit)
add("Gross margin %", [round((1 - COGS[i]) * 100, 1) for i in range(10)], fmt='0.0"%"')
ws.append([])
add("Opex")
for name, vals in opex.items():
    add(f"  {name}", vals)
add("Total opex", total_opex)
ws.append([])
add("EBITDA", ebitda)
add("EBITDA margin %", ebitda_margin, fmt='0.0"%"')
add("Net income / (loss)", ebitda)

# ---- Balance Sheet ----
ws2 = wb.create_sheet("Balance Sheet")
ws2.append(["Project TRINITY — Balance Sheet (₹ crore, illustrative)"])
ws2["A1"].font = NAVYAML2 = Font(bold=True, size=12)
ws2["A1"].font = NAVY
ws2.append([])
def render2(label, vals):
    ws2.append([label] + [v for v in vals])
    for c in range(2, 12):
        ws2.cell(row=ws2.max_row, column=c).number_format = '₹#,##0.00'
render2("Cash & equivalents", [round(max(total_revenue[i] * 0.4, 2), 2) for i in range(10)])
render2("Receivables", [round(total_revenue[i] * 0.1, 2) for i in range(10)])
render2("Capex asset (GPU cluster)", [round(0.5 * max(i - 4, 0) * 15, 2) for i in range(10)])
render2("Total assets", [round(max(total_revenue[i] * 0.4, 2) + total_revenue[i] * 0.1 + max(i - 4, 0) * 7.5, 2) for i in range(10)])
ws2.append([])
render2("Equity raised (cumulative)", [round(sum([2, 25, 100, 300, 700][:min(i + 1, 5)]), 2) for i in range(10)])
render2("Retained earnings", [round(sum(ebitda[:i + 1]) * 0.6, 2) for i in range(10)])
render2("Total L&E", [round(sum([2, 25, 100, 300, 700][:min(i + 1, 5)]) + sum(ebitda[:i + 1]) * 0.6, 2) for i in range(10)])

# ---- Cash Flow ----
ws3 = wb.create_sheet("Cash Flow")
ws3.append(["Project TRINITY — Cash Flow (₹ crore, illustrative)"])
ws3["A1"].font = NAVY
def render3(label, vals):
    ws3.append([label] + [v for v in vals])
    for c in range(2, 12):
        ws3.cell(row=ws3.max_row, column=c).number_format = '₹#,##0.00'
render3("Operating (≈ EBITDA)", ebitda)
capex = [round(0 if i < 5 else 15 * (i - 4), 2) for i in range(10)]
render3("Capex (cluster/campus)", capex)
render3("Free cash flow", [round(ebitda[i] - capex[i], 2) for i in range(10)])
injections = [2, 25, 100, 300, 500, 0, 0, 0, 0, 0]
render3("Funding raised (phased)", injections)
cash = 0
ending = []
for i in range(10):
    cash += ebitda[i] - capex[i] + injections[i]
    ending.append(round(cash, 2))
render3("Ending cash", ending)

wb.save(OUT)
print(f"SAVED {OUT}")